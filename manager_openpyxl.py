import openpyxl
from openpyxl import load_workbook
import pandas as pd
import xlsxwriter
import re

#-----Handle of file------------
# Load de file
file = r"C:\Users\Adila\Desktop\csv"
# Open de excel file
book_excel = openpyxl.load_workbook(file)
# List the excel sheet
name_sheet = book_excel.sheetnames
# Load the name in the first sheet
sheet = book_excel[name_sheet[0]]

#------Delete the first 2 rows-----------
# celdas = sheet.drop[1,2]
# sheet.drop(range(1, 2),axis=0)

# Calculate the last column and insert new column
last_column = sheet.max_column
new_column_platform = last_column + 1
new_column_status = new_column_platform + 1
last_row = sheet.max_row

#------Work to add 2 columns with fragmented file name---------
text = file.split('-')
aux_status = text[len(text) - 1].split('.')
aux_platform = text[len(text) - 2].split('\\')
status = aux_status[0]
platform = aux_platform[len(aux_platform) - 1]

for row in range(1, last_row):
    sheet.cell(column=new_column_platform, row=row).value = platform
    sheet.cell(column=new_column_status, row=row).value = status

#-------Save in csv format--------------------------------
new_nome = str(platform + "-" + status + ".csv")
save_doc = pd.DataFrame(sheet)
new_path = r'C:\csv' + new_nome
save_doc.to_csv(new_path)
book_excel.save(new_nome)

